import os
import csv
import pyttsx3
import datetime
from termcolor import colored




import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def format_csv_to_excel(input_name):
    # Determine if input_name is a file or a folder
    if os.path.isfile(input_name) and input_name.endswith('.csv'):
        # Process a single CSV file
        format_single_csv(input_name)
    elif os.path.isdir(input_name):
        # Process all CSV files in the folder
        for filename in os.listdir(input_name):
            if filename.endswith('.csv'):
                format_single_csv(os.path.join(input_name, filename))
    else:
        print("Invalid input. Please provide a CSV file or a folder containing CSV files.")

def format_single_csv(csv_file):
    # Load the CSV file into a DataFrame
    df = pd.read_csv(csv_file)

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    sheet = workbook.active

    # Write the DataFrame to the worksheet
    for r_idx, row in df.iterrows():
        for c_idx, value in enumerate(row):
            cell = sheet.cell(row=r_idx + 1, column=c_idx + 1, value=value)
            cell.alignment = Alignment(wrap_text=True)  # Enable text wrap

    # Adjust column widths
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2)  # Add some extra space
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook to an Excel file
    output_file = os.path.splitext(csv_file)[0] + '.xlsx'
    workbook.save(output_file)
    print(f"Formatted file saved as: {output_file}")

# Example usage:
# format_csv_to_excel('your_file.csv')  # For a single file
# format_csv_to_excel('your_folder')      # For all CSV files in a folder
















def speak(text):
    engine = pyttsx3.init("sapi5")
    voice = engine.getProperty("voices")
    engine.setProperty("voice", voice[1].id)
    engine.say(text)
    engine.runAndWait()

def check_folder():
    path = "Library"
    if not os.path.exists(path):
        os.makedirs(path)  # Create the directory and any necessary parent directories

def add_book():
    while True:
        speak("Enter the code number")
        code_no = input(colored("\n Enter the code number (0 to exit) : ","yellow")).lower()
        
        if code_no == "0":
            speak("Sorry! addition of book denied.")
            print(colored("\nSorry! addition of book denied.","red"))
            break
        
        speak("Enter the book name")
        book_name = input(colored("\n Enter the book name : ","white")).lower()
        
        speak("Enter the Author name")
        author_name = input(colored("\n Enter the Author name : ","yellow")).lower()
        
        speak("Enter the Publisher name")
        publisher_name = input(colored("\n Enter the Publisher name : ","white")).lower()
        
        speak("Enter the Cost of book")
        cost = input(colored("\n Enter the Cost of book: ","yellow")).lower()
        
        speak("Enter the Pages")
        pages = input(colored("\n Enter the Pages : ","white")).lower()
        
        speak("Enter the Year")
        year = input(colored("\n Enter the Year : ","yellow")).lower()
        
        speak("Enter the Quantity")
        quantity = input(colored("\n Enter the Quantity : ","white")).lower()
        
        speak("Enter the book category")
        book_category = input(colored("\n Enter the book category : ","yellow")).lower()
        
        speak("Enter the book place")
        book_place = input(colored("\n Enter the book place (i.e cupboard or shelve no.): ","white")).lower()
        
        check_folder()  # Ensure the 'Library' directory exists
        
        with open(f"Library\\{book_category}.csv", "w", newline='') as file:
            writer = csv.writer(file)
            
            writer.writerow([f"Code no."," "," ","Book Name"," "," ","Author Name"," "," ","Publisher Name"," "," ","Cost"," "," ","Pages"," "," ","Year"," "," ","Quantity"," "," ","Book Place"])
            writer.writerow([code_no," "," ",book_name," "," ",author_name," "," ",publisher_name," "," ",cost," "," ",pages," "," ",year," "," ",quantity," "," ",book_place]) 
            
            speak(f"Successfully added {book_name} in {book_category} category.")
            book_name = colored(f"{book_name}", "magenta")
            book_category = colored(f"{book_category}", "magenta")
            print(colored(f"\nSuccessfully added {book_name}","green"),end="")
            print(colored(f" in {book_category}","green"),end="")
            print(colored(f" category.","green"),end="")

def find_book():
    while True:
        speak("Enter the book information.")
        book_information = input(colored("\n Enter the book information (0 to exit): ","cyan")).lower()
        if book_information=="0":
            break
        else:
            for filename in os.listdir("Library"):
                if filename.endswith('.csv'):
                    with open(f"Library\\{filename}", "r", newline='') as file:
                        reader = csv.reader(file)
                        whole_list=list(reader)
                    
                    if not os.path.exists(f"Library\\{book_information}.csv"):
                        for filename in os.listdir("Library"):
                            if filename.endswith('.csv'):
                                print(f"Book information found in file: {filename}")
                                file_path = os.path.join("Library", filename)
                                with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
                                    reader = csv.reader(csvfile)
                                    for row in reader:
                                        if book_information in row:
                                            for item in row:
                                                # Only color the output, not the filename
                                                print(colored(item, "green"), end=" ")

                    
                    if os.path.exists(f"Library\\{book_information}.csv"):
                            f=1
                            speak(f"Here are some results for {book_information} category.")
                            for items in whole_list:
                                print(" ")
                                for value in items:
                                    print (colored(value,"magenta") , end=" ") if f==1 else print (colored(value,"cyan") , end=" ")
                            f+=1

def registration():
    def read_registration():
        with open("Registration.csv", "r", newline='') as file:
                reader = csv.reader(file)
                f=1
                for row in reader:
                    print(" ")
                    for items in row:
                        print (colored(items,"magenta"), end= " ") if f==1 else print(colored(items,"cyan"), end= " ")
                    f+=1
    
    while True:
        ask = input(colored("Enter\n0 to exit\n1 for ADD Registration\n2 for Read existing Registration\n3 for Delete Registration: ","cyan"))
        if ask == "0":
            break
        
        elif ask == "1":
            while True:
                speak("Enter the person's name (who is taking book)")
                person_name = input(colored("Enter the person's name (who is taking book) (OR 0 to exit): ","yellow")).lower()
                
                if person_name == "0":
                    speak("Sorry! addition of registration denied.")
                    print(colored("\nSorry! addition of registration denied.","red"))
                    break
                
                speak("Enter the book's name")
                book_name = input(colored("Enter the book's name : ","white")).lower()
                
                speak("Enter the category")
                book_category = input(colored("Enter the category: ","yellow")).lower()
                
                speak("Enter the time limit for which you are taking the book such as days")
                time_limit = input(colored(" Enter the time limit for which you are taking the book i.e days: ","white")).lower()
                
                current_datetime = datetime.datetime.now()
                time = current_datetime.strftime("%H:%M:%S")
                date = current_datetime.strftime("%d-%m-%Y")
                month = current_datetime.strftime("%B")
                year = current_datetime.strftime("%Y")
                if not os.path.exists("Registration.csv"):
                    with open("Registration.csv", "w", newline='') as file:
                        writer = csv.writer(file)
                        writer.writerow(["Person Name" ," "," ", "Time Limit" ," "," ", "Book Name"," " ," ", "Book Category"," " ," ", "Time"," " ," ", "Date"])
                        writer.writerow([f"{person_name}"," "," ", f"{time_limit}"," "," ", f"{book_name}"," "," ", f"{book_category}"," "," ", f"{time}"," "," ", f"{date},{month}, {year}"])
                else:
                    with open("Registration.csv", "a", newline='') as file:
                        writer = csv.writer(file)
                        writer.writerow([person_name, time_limit, book_name, book_category, time, f"{date}, {month}, {year}"])
                
                speak("Successfully Registered.")
                print(colored("Successfully Registered.","green"))
        
        elif ask == "2":
            read_registration()
        
        elif ask == "3":
            while True:
                speak("Enter the person's name (who is taking book)")
                person_name = input(colored("Enter the person's name (who is taking book) (OR 0 to exit): ","yellow")).lower()
                
                if person_name=='0':
                    break
                
                speak("Enter the book's name")
                book_name = input(colored("Enter the book's name : ","white")).lower()
                
                with open("Registration.csv", "r", newline='') as file:
                    reader = csv.reader(file)
                    lines = list(reader)
                
                filtered_lines = [item for item in lines if person_name not in item and book_name not in item]
                
                with open("Registration.csv", "w", newline='') as file:
                    writer = csv.writer(file)
                    writer.writerows(filtered_lines)

while True:
    inp = input(colored("\nEnter\n0 to exit\n1 to Add book\n2 to Find book\n3 to Register: ","blue"))
    if inp == "1":
        add_book()
    elif inp == "2":
        find_book()
    elif inp == "3":
        registration()
    elif inp == "0":
        speak("Have a nice day.")
        break

format_csv_to_excel("Registration.csv")
format_csv_to_excel("Library")

input("Enter to exit . . .")